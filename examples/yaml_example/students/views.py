import csv
import json

import yaml
from openpyxl import load_workbook
from django.shortcuts import render


def student_list(request):

    data_file_extension = request.GET.get('data_file_extension', 'yaml')
    file_path = f'media/students.{data_file_extension}'

    def get_default_students_list(file_path):
        with open(file_path, encoding='utf-8') as file:
            students = yaml.safe_load(file)

        return students

    match data_file_extension:
        case 'yaml':
            students = get_default_students_list(file_path)

        case 'csv':
            with open(file_path, 'r', encoding='utf-8') as file:
                csv_data = csv.DictReader(file)
                students = [row for row in csv_data]

        case 'json':
            with open(file_path, 'r', encoding='utf-8') as file:
                students = json.load(file)

        case 'xlsx':
            workbook = load_workbook(filename=file_path)
            worksheet = workbook.active

            skip_first_row = True
            min_row = 2 if skip_first_row else 1
            students = [
                {
                    'name': row[0],
                    'lastname': row[1],
                    'age': row[2],
                    'faculty': row[3],
                    'year': row[4],
                    'average': row[5]
                }
                for row in worksheet.iter_rows(values_only=True, min_row=min_row)
            ]

        case _:
            students = get_default_students_list('media/students.yaml')

    return render(request, 'student_list.html', {'students': students})
